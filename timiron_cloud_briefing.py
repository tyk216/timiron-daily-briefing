"""
timiron_cloud_briefing.py — Timiron Daily Briefing (v4)
Runs on GitHub Actions at 6 AM ET daily (10:00 UTC).
Pulls load log from cadiz.ops Outlook email via Microsoft Graph API.
Sends dark-themed HTML briefing + two Excel attachments via Gmail SMTP.

v4 changes:
  - Month-agnostic (works for any month, not hardcoded to March)
  - Uses $filter instead of $search for deterministic email results
  - Retry logic on all Graph API calls
  - Error notification email on failure
  - Dynamic week boundaries, YTD calculation, Excel row targeting
"""

import os, json, re, sys, shutil, smtplib, base64, io, tempfile, calendar, time, traceback
from datetime import date, timedelta, datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import requests
import pandas as pd
import openpyxl

# ════════════════════════════════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════════════════════════════════

MS_GRAPH_REFRESH_TOKEN = os.environ.get('MS_GRAPH_REFRESH_TOKEN', '')
MS_GRAPH_CLIENT_ID     = os.environ.get('MS_GRAPH_CLIENT_ID', '')
GMAIL_ADDRESS          = os.environ.get('GMAIL_ADDRESS', 'tyk216@gmail.com')
GMAIL_APP_PASS         = os.environ.get('GMAIL_APP_PASS', '')
RECIPIENTS             = os.environ.get('RECIPIENTS', 'tylerk@timironmp.com,robk@timirontrading.com').split(',')
QBT_TOKEN              = os.environ.get('QBT_TOKEN', '')

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(SCRIPT_DIR, "templates")

CARRIER_AVGS = {
    'Badlands':          224.6,
    'KAG':               188.9,
    'Prop Logistics':    181.5,
    'BD Oil':            188.2,
    '1st Choice Energy': 183.3,
}

# ════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ════════════════════════════════════════════════════════════════════════════

RAIL_CAP_DAILY    = 15000
PUMP_AVAIL_HRS    = 21           # 24 minus 3hr rail switch
FEB_2026_AVG      = 11200        # benchmark daily avg
FEB_2026_TOTAL    = 313600       # benchmark month total

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
TOKEN_URL  = "https://login.microsoftonline.com/common/oauth2/v2.0/token"

ACCESS_TOKEN = None

def rev_per_day(bbls):
    """Tiered revenue calculation: $1.30 first 5k, $0.95 next 5k, $0.75 above 10k."""
    return (min(bbls, 5000) * 1.30 + max(0, min(bbls - 5000, 5000)) * 0.95 + max(0, bbls - 10000) * 0.75)

def fmt_date(d):
    """Format date without zero-padding (cross-platform)."""
    return f"{d.strftime('%b')} {d.day}, {d.year}"

def fmt_date_short(d):
    """e.g. 'Mar 31'"""
    return f"{d.strftime('%b')} {d.day}"

def fmt_date_file(d):
    """e.g. '4-1-26' for filenames."""
    return f"{d.month}-{d.day}-{d.strftime('%y')}"

def month_name(d):
    """e.g. 'Apr 2026'"""
    return d.strftime('%b %Y')

# ════════════════════════════════════════════════════════════════════════════
# AUTH
# ════════════════════════════════════════════════════════════════════════════


def _persist_refresh_token(new_rt, env_path="/opt/briefing/briefing.env"):
    """Atomically rewrite MS_GRAPH_REFRESH_TOKEN line in briefing.env.

    Azure rotates refresh tokens on each use (offline_access, single-tenant MSA).
    If we don't persist the rotated token, the next run fails with AADSTS900144
    / malformed JWT. Called from get_access_token() on every successful refresh.
    """
    import os
    with open(env_path) as _f:
        lines = _f.readlines()
    out, found = [], False
    for ln in lines:
        if ln.startswith("MS_GRAPH_REFRESH_TOKEN="):
            out.append("MS_GRAPH_REFRESH_TOKEN=" + new_rt + "\n")
            found = True
        else:
            out.append(ln)
    if not found:
        out.append("MS_GRAPH_REFRESH_TOKEN=" + new_rt + "\n")
    tmp = env_path + ".tmp"
    with open(tmp, "w") as _f:
        _f.write("".join(out))
    os.chmod(tmp, 0o600)
    os.replace(tmp, env_path)
    # Update in-memory constant so the rest of this run sees the new value
    globals()["MS_GRAPH_REFRESH_TOKEN"] = new_rt


def get_access_token():
    global ACCESS_TOKEN
    for attempt in range(3):
        try:
            r = requests.post(TOKEN_URL, data={
                "client_id":     MS_GRAPH_CLIENT_ID,
                "grant_type":    "refresh_token",
                "refresh_token": MS_GRAPH_REFRESH_TOKEN,
                "scope":         "Mail.Read Files.Read.All offline_access",
            }, timeout=30)
            if r.ok:
                data = r.json()
                ACCESS_TOKEN = data["access_token"]
                new_rt = data.get("refresh_token")
                if new_rt and new_rt != MS_GRAPH_REFRESH_TOKEN:
                    try:
                        _persist_refresh_token(new_rt)
                        print("  ✓ Rotated refresh token persisted to briefing.env")
                    except Exception as _pe:
                        print(f"  ⚠ Rotated refresh token received but PERSIST FAILED: {_pe}")
                        print(f"  Manual update needed (first 20 chars): {new_rt[:20]}...")
                print("  Access token acquired.")
                return True
            print(f"  Token refresh attempt {attempt+1} failed: {r.status_code} {r.text[:200]}")
        except Exception as e:
            print(f"  Token refresh attempt {attempt+1} error: {e}")
        if attempt < 2:
            time.sleep(3)
    return False

def graph_headers():
    return {"Authorization": f"Bearer {ACCESS_TOKEN}", "Content-Type": "application/json"}

# ════════════════════════════════════════════════════════════════════════════
# GRAPH HELPERS — with retry and $filter support
# ════════════════════════════════════════════════════════════════════════════

def graph_get(url, params=None, retries=3):
    """GET request to Graph API with retry logic."""
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=graph_headers(), params=params, timeout=60)
            if r.ok:
                return r.json()
            if r.status_code == 429:  # throttled
                wait = int(r.headers.get('Retry-After', 10))
                print(f"  Throttled, waiting {wait}s...")
                time.sleep(wait)
                continue
            print(f"  Graph GET failed ({attempt+1}/{retries}): {r.status_code} {r.text[:200]}")
        except Exception as e:
            print(f"  Graph GET error ({attempt+1}/{retries}): {e}")
        if attempt < retries - 1:
            time.sleep(3)
    return None

def filter_emails(sender, since, until=None, top=10, select=None):
    """Query emails by sender and date range using $filter (no contains on subject — Graph doesn't support it on all mailboxes)."""
    url = f'{GRAPH_BASE}/me/messages'
    filt = f"from/emailAddress/address eq '{sender}' and receivedDateTime ge {since}"
    if until:
        filt += f" and receivedDateTime le {until}"
    params = {
        "$filter": filt,
        "$top": top,
        "$orderby": "receivedDateTime desc",
    }
    if select:
        params["$select"] = select
    data = graph_get(url, params)
    return data.get("value", []) if data else []

def search_emails(search_query, top=5):
    """Search emails using $search (relevance-based). Used as fallback."""
    url = f'{GRAPH_BASE}/me/messages'
    params = {
        "$search": f'"{search_query}"',
        "$top": top,
        "$select": "id,subject,from,body,receivedDateTime,hasAttachments",
    }
    data = graph_get(url, params)
    return data.get("value", []) if data else []

def get_attachments(message_id):
    url = f'{GRAPH_BASE}/me/messages/{message_id}/attachments'
    data = graph_get(url)
    return data.get("value", []) if data else []

def get_body_text(msg):
    body = msg.get("body", {})
    content = body.get("content", "")
    if body.get("contentType") == "html":
        content = re.sub(r'<br\s*/?>', '\n', content, flags=re.IGNORECASE)
        content = re.sub(r'<[^>]+>', '', content)
        content = re.sub(r'&nbsp;', ' ', content)
        content = re.sub(r'&#\d+;', '', content)
    return content.strip()

# ════════════════════════════════════════════════════════════════════════════
# FETCH CADIZ OPS DATA — switch times, carrier replies
# ════════════════════════════════════════════════════════════════════════════

def fetch_cadiz_ops(yesterday):
    """Search Outlook for switch times and carrier projections for yesterday."""
    today = yesterday + timedelta(days=1)
    yday_str = yesterday.strftime('%m.%d.%y')  # e.g. 03.31.26
    # Also handle single-digit month/day formats used by cadiz ops
    yday_str_alt = f"{yesterday.month}.{yesterday.day}.{yesterday.strftime('%y')}"  # e.g. 3.31.26

    result = {
        "date": yday_str,
        "switch_times": [],  # list of {start, end, duration}
        "updates": [],       # raw update email bodies
        "loaded_cars_out": 0,
        "empty_cars_in": 0,
        "maintenance_notes": [],
        "carrier_projections": {},
    }

    since = (yesterday - timedelta(days=1)).strftime('%Y-%m-%dT00:00:00Z')
    until = today.strftime('%Y-%m-%dT23:59:59Z')

    print("  Searching for UPDATE emails...")
    msgs = filter_emails('cadiz.ops@timirontrading.com', since, until, top=20,
                         select="id,subject,body,receivedDateTime,hasAttachments")

    # Filter UPDATE emails: check subject in Python (Graph $filter can't do contains on subject)
    for msg in msgs:
        subj = msg.get("subject", "")
        subj_upper = subj.upper()
        # Must have UPDATE in subject and yesterday's date
        if 'UPDATE' not in subj_upper:
            continue
        if yday_str not in subj and yday_str_alt not in subj:
            continue
        body = get_body_text(msg)
        result["updates"].append({"subject": subj, "body": body})

        # Extract switch times
        if re.search(r'(ON SITE|ARRIVED|HAS ARRIVED)', body, re.IGNORECASE):
            time_m = re.search(r'(\d{1,2}:\d{2}\s*[AP]M)\s+UPDATE', subj, re.IGNORECASE)
            if time_m:
                result["switch_times"].append({"start": time_m.group(1).strip(), "type": "start"})
        if re.search(r'(CLEARED|COMPLETED|RESUMING|NORMAL OPERATIONS)', body, re.IGNORECASE):
            time_m = re.search(r'(\d{1,2}:\d{2}\s*[apAP][mM])\s+[uU]', subj, re.IGNORECASE)
            if time_m:
                result["switch_times"].append({"end": time_m.group(1).strip(), "type": "end"})

        # Maintenance notes
        maint_keywords = ['pump', 'repair', 'replace', 'fix', 'broke', 'leak', 'down',
                          'out of service', 'maintenance', 'welding', 'valve', 'hose', 'motor']
        for line in body.split('\n'):
            line = line.strip()
            if line and len(line) > 10 and any(kw in line.lower() for kw in maint_keywords):
                result["maintenance_notes"].append(line[:200])

    # RAIL SWAP email (same cadiz.ops emails, check for RAIL in subject in Python)
    print("  Searching for RAIL SWAP email...")
    for msg in msgs:
        subj = msg.get("subject", "")
        if 'RAIL' not in subj.upper():
            continue
        if yday_str not in subj and yday_str_alt not in subj:
            continue
        body = get_body_text(msg)
        loaded_m = re.search(r'(\d+)\s+LOADED\s+CARS?\s+SENT', body, re.IGNORECASE)
        empty_m = re.search(r'(\d+)\s+EMPTY\s+CARS?\s+PUSHED', body, re.IGNORECASE)
        if loaded_m:
            result["loaded_cars_out"] = int(loaded_m.group(1))
        if empty_m:
            result["empty_cars_in"] = int(empty_m.group(1))

    # Carrier projections
    carrier_senders = {
        'Badlands': 'ohiodispatch@badlands-ngl.com',
        'KAG':     'bxi-bloomingdale@kagcentral.com',
    }
    for cname, sender in carrier_senders.items():
        print(f"  Searching for {cname} carrier reply...")
        today_start = today.strftime('%Y-%m-%dT00:00:00Z')
        c_msgs = filter_emails(sender, today_start, top=3,
                               select="id,subject,body,receivedDateTime")
        trucks = 0
        note = "No response"
        for msg in c_msgs:
            body = get_body_text(msg)
            truck_m = re.search(r'(\d+)\s+(?:planned|trucks?|loads?|scheduled)', body, re.IGNORECASE)
            if not truck_m:
                truck_m = re.search(r'(?:have|running|sending|doing)\s+(\d+)', body, re.IGNORECASE)
            if truck_m:
                trucks = int(truck_m.group(1))
                note = ""
            break
        avg = CARRIER_AVGS.get(cname, 190)
        result["carrier_projections"][cname] = {
            "trucks": trucks, "proj_bbls": round(trucks * avg), "note": note
        }

    for cname in ['Prop Logistics', 'BD Oil', '1st Choice Energy']:
        if cname not in result["carrier_projections"]:
            result["carrier_projections"][cname] = {"trucks": 0, "proj_bbls": 0, "note": "No response"}

    return result

# ════════════════════════════════════════════════════════════════════════════
# FETCH MASTER LOAD LOG — via $filter for reliable date-ordered results
# ════════════════════════════════════════════════════════════════════════════

def fetch_load_log_excel(yesterday):
    """Find the most recent cadiz.ops LOGS email and download the Master Load Log attachment."""
    print("  Searching for LOGS email from cadiz.ops...")

    # Filter by sender + date range, then check subject in Python
    since = (yesterday - timedelta(days=2)).strftime('%Y-%m-%dT00:00:00Z')
    msgs = filter_emails('cadiz.ops@timirontrading.com', since, top=10,
                         select="id,subject,receivedDateTime,hasAttachments")
    # Only keep messages with LOGS in subject
    msgs = [m for m in msgs if 'LOGS' in m.get('subject', '').upper()]

    if not msgs:
        # Fallback to $search
        print("  $filter returned nothing, falling back to $search...")
        msgs = search_emails("from:cadiz.ops subject:LOGS")

    for msg in msgs:
        if not msg.get("hasAttachments", True):
            continue
        attachments = get_attachments(msg["id"])
        for att in attachments:
            name = att.get("name", "")
            # Match: "MASTER COPY - 1Q 2026 ..." or "MASTER COPY - FEB MASTER LOAD LOG ..."
            if name.lower().endswith(('.xlsx', '.xls')) and ('load log' in name.lower() or ('master copy' in name.lower() and 'railcar' not in name.lower())):
                content_bytes = att.get("contentBytes")
                if content_bytes:
                    excel_bytes = base64.b64decode(content_bytes)
                    print(f"    Found: {name} ({len(excel_bytes):,} bytes) from {msg.get('subject','')}")
                    return excel_bytes, name
    print("  ERROR: Could not find Master Load Log Excel attachment")
    return None, None

# ════════════════════════════════════════════════════════════════════════════
# PARSE LOAD LOG — month-agnostic
# ════════════════════════════════════════════════════════════════════════════

def parse_load_log(excel_bytes, yesterday):
    """Parse the Master Load Log. Returns dict with all KPIs."""
    df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name='Master_Load_Log', header=0)
    df['Date'] = pd.to_datetime(df['Date']).dt.date
    df['BOL_prefix'] = df['Timiron BOL#'].astype(str).str[:3]
    df['Metered'] = pd.to_numeric(df['Timiron Metered bbls.'], errors='coerce').fillna(0)

    def to_mins(t):
        """Convert pump time to minutes. Handles time objects, strings, and fractions."""
        if pd.isna(t) or t is None:
            return 0
        if hasattr(t, 'hour'):  # datetime.time object
            return t.hour * 60 + t.minute + t.second / 60
        if isinstance(t, (int, float)):
            return t * 24 * 60  # fraction of day
        try:
            s = str(t)
            p = s.split(':')
            return int(p[0]) * 60 + int(p[1])
        except:
            return 0

    df['pump_mins'] = df['Pump Time'].apply(to_mins)

    # Determine the reporting month from yesterday's date
    month_start = yesterday.replace(day=1)
    _, days_in_month = calendar.monthrange(yesterday.year, yesterday.month)
    month_name_str = yesterday.strftime('%B %Y')
    month_abbr = yesterday.strftime('%b')
    prev_month_name = (month_start - timedelta(days=1)).strftime('%b %Y')

    # Filter to current month data
    mtd_data = df[(df['Date'] >= month_start) & (df['Date'] <= yesterday)]
    if mtd_data.empty:
        raise ValueError(f"No data for {month_name_str} in load log.")

    # Yesterday's data
    yday = df[df['Date'] == yesterday]
    yday_count = len(yday)
    print(f"  Yesterday: {yesterday}  ({yday_count} loads)")

    mtd_days = sorted(mtd_data['Date'].unique())
    print(f"  MTD: {len(mtd_days)} days  ({min(mtd_days)} -- {max(mtd_days)})")

    # Pump utilization (yesterday)
    pump_map = {'111': 'P-101', '222': 'P-102', '333': 'P-103'}
    pump_ute = {}
    for prefix, pname in pump_map.items():
        p = yday[yday['BOL_prefix'] == prefix]
        splits = p[p['Split Load'].astype(str).str.contains('Split #2', na=False)]
        non_split = p[~p['Split Load'].astype(str).str.contains('Split #2', na=False)]
        runtime = p['pump_mins'].sum() / 60
        bbls = p['Metered'].sum()
        pump_ute[pname] = {
            'loads': len(non_split), 'splits': len(splits),
            'runtime': round(runtime, 2), 'ute': round(runtime / PUMP_AVAIL_HRS * 100, 1),
            'bbls': round(bbls, 2),
            'bbl_hr': round(bbls / runtime, 0) if runtime > 0 else 0
        }
    combined_rt = sum(v['runtime'] for v in pump_ute.values())

    # MTD totals — run rate uses only COMPLETED days (exclude any partial today data)
    is_split2 = lambda s: pd.notna(s) and 'split' in str(s).lower() and '2' in str(s)
    mtd_completed = mtd_data[mtd_data['Date'] <= yesterday]  # only full days through yesterday
    mtd_no_split = mtd_completed[~mtd_completed['Split Load'].apply(is_split2)]
    daily_trucks = mtd_no_split.groupby('Date').size()
    daily_bbls = mtd_completed.groupby('Date')['Metered'].sum()
    total_bbls = daily_bbls.sum()
    total_trucks = daily_trucks.sum()
    days_actual = len(daily_bbls)
    days_remain = days_in_month - yesterday.day
    avg_bbls = total_bbls / days_actual if days_actual > 0 else 0
    avg_trucks = total_trucks / days_actual if days_actual > 0 else 0
    proj_bbls = total_bbls + avg_bbls * days_remain
    proj_trucks = total_trucks + avg_trucks * days_remain
    proj_rev = rev_per_day(avg_bbls) * days_in_month
    rail_cap = avg_bbls / RAIL_CAP_DAILY

    # MTD pump hours
    p101_hrs = mtd_data[mtd_data['BOL_prefix'] == '111']['pump_mins'].sum() / 60
    p102_hrs = mtd_data[mtd_data['BOL_prefix'] == '222']['pump_mins'].sum() / 60
    p103_hrs = mtd_data[mtd_data['BOL_prefix'] == '333']['pump_mins'].sum() / 60
    total_pump_hrs = p101_hrs + p102_hrs + p103_hrs
    pump_ute_combined = total_pump_hrs / (days_actual * PUMP_AVAIL_HRS * 3) if days_actual > 0 else 0

    print(f"  MTD BBLs:  {total_bbls:,.2f}  avg {avg_bbls:,.1f}/day")
    print(f"  Projected: {proj_bbls:,.0f} BBLs | {proj_trucks:,.0f} trucks")
    print(f"  Pump Ute:  P-101 {pump_ute['P-101']['ute']}%  P-102 {pump_ute['P-102']['ute']}%  P-103 {pump_ute['P-103']['ute']}%  Combined {pump_ute_combined*100:.1f}%")

    # Carrier actuals (yesterday)
    carrier_name_map = {'BD OIL': 'BD Oil'}
    carrier_actuals = {}
    if 'Carrier' in df.columns:
        yday_no_split = yday[~yday['Split Load'].apply(is_split2)]
        # Guard: empty yday (no loads yet for target date) drops columns in some
        # pandas versions and raises KeyError on groupby. Skip gracefully.
        if not yday_no_split.empty and 'Carrier' in yday_no_split.columns:
            for carrier_name, grp in yday_no_split.groupby('Carrier'):
                normalized = carrier_name_map.get(carrier_name, carrier_name)
                carrier_actuals[normalized] = {
                    'trucks': len(grp),
                    'bbls': round(yday[yday['Carrier'] == carrier_name]['Metered'].sum(), 1),
                }
        else:
            print(f'  Carrier actuals: skipped (yday rows={len(yday)}, yday_no_split rows={len(yday_no_split)}) — likely load log not yet updated for target date')

    # API Gravity and BSW
    avg_api = 0
    avg_bsw = 0
    if 'Timiron API Gravity  Meter' in yday.columns:
        vals = pd.to_numeric(yday['Timiron API Gravity  Meter'], errors='coerce').dropna()
        avg_api = round(vals.mean(), 2) if len(vals) > 0 else 0
    if 'BSW%' in yday.columns:
        vals = pd.to_numeric(yday['BSW%'], errors='coerce').dropna()
        avg_bsw = round(vals.mean(), 5) if len(vals) > 0 else 0

    # Daily data for 5-day trend
    daily_data = []
    for d_date in sorted(mtd_data['Date'].unique()):
        day_df = mtd_data[mtd_data['Date'] == d_date]
        day_no_split = day_df[~day_df['Split Load'].apply(is_split2)]
        daily_data.append({
            'date': d_date,
            'bbls': round(day_df['Metered'].sum(), 1),
            'trucks': len(day_no_split),
            'day_name': d_date.strftime('%a'),
        })

    # Weekly breakdown (dynamic for any month)
    weekly_data = []
    wk_num = 1
    wk_start_day = 1
    while wk_start_day <= days_in_month:
        wk_end_day = min(wk_start_day + 6, days_in_month)
        w_start = date(yesterday.year, yesterday.month, wk_start_day)
        w_end = date(yesterday.year, yesterday.month, wk_end_day)
        w_days = [dd for dd in daily_data if w_start <= dd['date'] <= w_end]
        if w_days:
            w_bbls = sum(dd['bbls'] for dd in w_days)
            w_trucks = sum(dd['trucks'] for dd in w_days)
            w_avg = w_bbls / len(w_days)
            w_bpt = w_bbls / w_trucks if w_trucks > 0 else 0
            weekly_data.append({
                'label': f"Wk{wk_num} ({month_abbr} {wk_start_day}-{wk_end_day})",
                'bbls': round(w_bbls, 1), 'trucks': w_trucks,
                'days': len(w_days), 'avg_bbls': round(w_avg, 1), 'avg_bpt': round(w_bpt, 1),
            })
        wk_start_day = wk_end_day + 1
        wk_num += 1

    # Compute fixed cost estimate (scale from known Feb cost)
    feb_cost = 234498.18
    fixed_cost_est = feb_cost * (days_in_month / 28)  # rough scaling
    ebitda = proj_rev - fixed_cost_est

    return dict(
        yesterday_date=yesterday, month_start=month_start,
        month_name=month_name_str, month_abbr=month_abbr,
        days_in_month=days_in_month,
        days_actual=days_actual, days_remain=days_remain,
        total_bbls=round(total_bbls, 2), total_trucks=int(total_trucks),
        avg_bbls=round(avg_bbls, 1), avg_trucks=round(avg_trucks, 1),
        proj_bbls=round(proj_bbls, 0), proj_trucks=round(proj_trucks, 0),
        proj_rev=round(proj_rev, 2), ebitda=round(ebitda, 2),
        fixed_cost=round(fixed_cost_est, 2),
        rail_cap=round(rail_cap, 6), pump_ute=pump_ute,
        pump_ute_combined=round(pump_ute_combined, 3),
        p101_hrs=round(p101_hrs, 2), p102_hrs=round(p102_hrs, 2), p103_hrs=round(p103_hrs, 2),
        total_pump_hrs=round(total_pump_hrs, 2),
        carrier_actuals=carrier_actuals,
        avg_api_gravity=avg_api, avg_bsw=avg_bsw,
        daily_data=daily_data, weekly_data=weekly_data,
    )

# ════════════════════════════════════════════════════════════════════════════
# UPDATE EXCEL FILES
# ════════════════════════════════════════════════════════════════════════════

def safe_write(ws, row, col, val):
    cell = ws.cell(row, col)
    if cell.__class__.__name__ != 'MergedCell':
        cell.value = val

def find_template(name_fragment):
    exact = os.path.join(TEMPLATE_DIR, f"Timiron_{name_fragment}.xlsx")
    if os.path.exists(exact):
        print(f"  Template: Timiron_{name_fragment}.xlsx")
        return exact
    raise FileNotFoundError(f"Template not found: {exact}")

def find_month_row(ws, month_str, default_row=16):
    """Find the row containing the month string (e.g. 'Mar 2026')."""
    for r in range(4, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and month_str in str(val):
            return r
    return default_row

def update_dashboard(template_path, d, output_path, crew_hours=None):
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    # Operations Dashboard — v15 column layout:
    # C1:Month C2:Headcount C3:Revenue C4:Cost C5:Net Income
    # C6:Total Bbls C7:Bbls/Day C8:Trucks/Day C9:Total Trucks
    # C10:Rev/Bbl C11:Cost/Bbl C12:Net/Bbl C13:Pump Util% C14:Rail Cap C15:Net/Head
    ws = wb['Operations Dashboard']
    r = find_month_row(ws, d['month_abbr'])
    dt = fmt_date_short(d['yesterday_date'])
    suffix = '\u2020' if d['days_remain'] > 0 else ''
    bbls_val = d['proj_bbls'] if d['days_remain'] > 0 else d['total_bbls']
    trucks_val = d['proj_trucks'] if d['days_remain'] > 0 else d['total_trucks']
    headcount = ws.cell(r, 2).value or 15  # preserve existing headcount

    ws.cell(r, 1).value = f"{d['month_name']}{suffix}"
    ws.cell(r, 3).value = d['proj_rev']                                                     # C3: Revenue
    ws.cell(r, 5).value = d['ebitda']                                                        # C5: Net Income
    ws.cell(r, 6).value = round(bbls_val)                                                    # C6: Total Barrels
    ws.cell(r, 7).value = d['avg_bbls']                                                      # C7: Bbls/Day
    ws.cell(r, 8).value = d['avg_trucks']                                                    # C8: Trucks/Day
    ws.cell(r, 9).value = round(trucks_val)                                                  # C9: Total Trucks
    ws.cell(r, 10).value = round(d['proj_rev'] / bbls_val, 2) if bbls_val > 0 else 0        # C10: Rev/Bbl
    ws.cell(r, 11).value = round(d['fixed_cost'] / bbls_val, 3) if bbls_val > 0 else 0      # C11: Cost/Bbl
    ws.cell(r, 12).value = round(d['ebitda'] / bbls_val, 3) if bbls_val > 0 else 0          # C12: Net/Bbl
    ws.cell(r, 13).value = d['pump_ute_combined']                                            # C13: Pump Util%
    ws.cell(r, 14).value = d['rail_cap']                                                     # C14: Rail Cap
    net_per_head = round(d['ebitda'] / headcount) if headcount else 0
    ws.cell(r, 15).value = net_per_head                                                      # C15: Net/Head

    # Pump Runtime — C180 fix 2026-04-13:
    # Use calendar-hours avail (days × 24) for Down Hrs, Total Hrs, and Util% so
    # Run + Down = Total Hrs column for partial AND full months. Historical rows
    # (Feb 2026 etc.) already use full-month calendar hours (28×24=672); this
    # write matches that convention. Partial-month rows (April 2026†) show
    # days_actual×24 = MTD calendar hours, clearly signalled by the † suffix.
    pr = wb['Pump Runtime']
    pr_row = find_month_row(pr, d['month_abbr'], 15)
    calendar_avail = d['days_actual'] * 24   # per-pump MTD calendar hours

    pr.cell(pr_row, 1).value = f"{d['month_name']}{suffix}"
    pr.cell(pr_row, 2).value = calendar_avail  # MTD calendar hrs (was full-month; caused Run+Down≠Total on partials)
    pr.cell(pr_row, 3).value = round(d['p101_hrs'], 2)
    pr.cell(pr_row, 4).value = round(calendar_avail - d['p101_hrs'], 2)
    pr.cell(pr_row, 5).value = round(d['p101_hrs'] / calendar_avail, 3) if calendar_avail > 0 else 0
    pr.cell(pr_row, 6).value = round(d['p102_hrs'], 2)
    pr.cell(pr_row, 7).value = round(calendar_avail - d['p102_hrs'], 2)
    pr.cell(pr_row, 8).value = round(d['p102_hrs'] / calendar_avail, 3) if calendar_avail > 0 else 0
    pr.cell(pr_row, 9).value = round(d['p103_hrs'], 2)
    pr.cell(pr_row, 10).value = round(calendar_avail - d['p103_hrs'], 2)
    pr.cell(pr_row, 11).value = round(d['p103_hrs'] / calendar_avail, 3) if calendar_avail > 0 else 0
    partial = f"Partial month thru {dt}" if d['days_remain'] > 0 else "Full month actuals"
    pr.cell(pr_row, 12).value = f"All 3 pumps active \u00b7 {partial}"

    # Headcount & Roster — WTD hours from QBT
    if crew_hours and 'Headcount & Roster' in wb.sheetnames:
        hr = wb['Headcount & Roster']
        # Build lookup: normalize name -> hours
        hrs_lookup = {}
        for ch in crew_hours:
            hrs_lookup[ch["name"].lower()] = ch["total"]

        # Map dashboard names (rows 10-24, col B) to QBT data (col D = WTD)
        for row in range(10, 25):
            dash_name = hr.cell(row=row, column=2).value
            if not dash_name:
                continue
            # Normalize: "Shawn Osborne Jr." -> match "Shawn Osborn Jr.", etc.
            dn = dash_name.lower().replace("osborne", "osborn").replace(" (oz)", "")
            matched = hrs_lookup.get(dn)
            if matched is not None:
                hr.cell(row=row, column=4, value=round(matched, 1))

        # Update header
        from zoneinfo import ZoneInfo
        ET_tz = ZoneInfo("America/New_York")
        now_et = datetime.now(ET_tz)
        today_et = now_et.date()
        monday = today_et - timedelta(days=today_et.weekday())
        hr.cell(row=9, column=4, value=f"WTD\n{fmt_date_short(monday)}-{fmt_date_short(today_et)}")
        hr.cell(row=27, column=1, value=f"WTD = Week to date through {fmt_date_short(today_et)} (includes active night shift). Avg based on 4 full weeks only.")

    # === XLSX REVAMP 2026-04-13 — unify April forecast across all tabs ===
    # Before: Ops Dashboard R16 (live) + Apr P&L Forecast B5 (static pre-month) +
    # Apr Weekly R4-9 (static pre-month) — three different Aprils. Now: one truth.
    stamp = f"Live update \u2014 {fmt_date_short(d['yesterday_date'])} data, {d['days_actual']}-day run rate"

    # --- Apr 2026 P&L Forecast tab: overwrite Services revenue with live projection ---
    pl_sheet_name = None
    for candidate in (f"{d['month_abbr']} 2026 P&L Forecast", "Apr 2026 P&L Forecast"):
        if candidate in wb.sheetnames:
            pl_sheet_name = candidate
            break
    if pl_sheet_name:
        pl = wb[pl_sheet_name]
        safe_write(pl, 5, 2, round(d['proj_rev'], 2))
        safe_write(pl, 2, 2, stamp)
        for r in range(67, min(75, pl.max_row + 1)):
            val = pl.cell(r, 1).value
            if val and isinstance(val, str) and val.startswith("Apr 2026"):
                safe_write(pl, r, 1, (
                    f"Apr 2026 = Live projection based on {d['days_actual']} days of actuals "
                    f"(run rate {d['avg_bbls']:,.0f} bbls/day \u2192 {d['proj_bbls']:,.0f} bbls projected total). "
                    f"Revenue cell is live-updated daily; expense lines scaled from Mar 2026 baseline."
                ))

    # --- Apr 2026 Weekly tab: rewrite per-week rows with actual/forecast split ---
    wk_sheet_name = None
    for candidate in (f"{d['month_abbr']} 2026 Weekly", "Apr 2026 Weekly"):
        if candidate in wb.sheetnames:
            wk_sheet_name = candidate
            break
    if wk_sheet_name:
        wk = wb[wk_sheet_name]
        days_in_m = d['days_in_month']
        boundaries = [(1, 7), (8, 14), (15, 21), (22, 28), (29, days_in_m)]
        daily = d.get('daily_data', []) or []
        by_day = {dd['date'].day: dd for dd in daily}
        yday_dom = d['yesterday_date'].day
        run_rate = d['avg_bbls']
        avg_trucks_day = d.get('avg_trucks', 0)
        rev_per_bbl_live = d['proj_rev'] / d['proj_bbls'] if d['proj_bbls'] else 0
        cost_per_day = d['fixed_cost'] / days_in_m if days_in_m else 0

        for i, (start, end) in enumerate(boundaries):
            if start > days_in_m:
                break
            row = 4 + i
            days_count = end - start + 1
            actual_bbls = sum(by_day[doy]['bbls'] for doy in range(start, end + 1) if doy in by_day)
            actual_trucks = sum(by_day[doy]['trucks'] for doy in range(start, end + 1) if doy in by_day)
            actual_days = sum(1 for doy in range(start, end + 1) if doy in by_day)
            forecast_days = max(0, days_count - actual_days)
            forecast_bbls = forecast_days * run_rate
            week_bbls = actual_bbls + forecast_bbls
            week_rev = week_bbls * rev_per_bbl_live
            week_exp = days_count * cost_per_day
            if end <= yday_dom:
                status = "COMPLETE"
            elif start > yday_dom:
                status = "FORECAST"
            else:
                status = f"IN PROGRESS ({actual_days}/{days_count})"
            safe_write(wk, row, 1, f"Week {i + 1}")
            safe_write(wk, row, 2, f"{d['month_abbr']} {start}\u2013{end}")
            safe_write(wk, row, 3, days_count)
            safe_write(wk, row, 4, round(week_bbls))
            safe_write(wk, row, 5, round(week_rev, 2))
            safe_write(wk, row, 6, round(week_exp, 2))
            safe_write(wk, row, 7, round(week_rev - week_exp, 2))
            safe_write(wk, row, 8, status)

        weeks_written = sum(1 for (s, e) in boundaries if s <= days_in_m)
        for extra_row in range(4 + weeks_written, 9):
            for c in range(1, 9):
                safe_write(wk, extra_row, c, None)

        safe_write(wk, 9, 1, "TOTAL")
        safe_write(wk, 9, 2, f"{d['month_abbr']} 1\u2013{days_in_m}")
        safe_write(wk, 9, 3, days_in_m)
        safe_write(wk, 9, 4, round(d['proj_bbls']))
        last_data_row = 4 + weeks_written - 1
        safe_write(wk, 9, 5, f"=SUM(E4:E{last_data_row})")
        safe_write(wk, 9, 6, f"=SUM(F4:F{last_data_row})")
        safe_write(wk, 9, 7, "=E9-F9")
        if wk.max_row >= 11:
            safe_write(wk, 11, 1, f"RUN RATE (LIVE): {run_rate:,.0f} bbls/day \u00b7 {stamp}")

    wb.save(output_path)
    print(f"  Dashboard saved: {os.path.basename(output_path)}")

def update_external_report(template_path, d, output_path):
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    dt = fmt_date_short(d['yesterday_date'])
    suffix = '\u2020' if d['days_remain'] > 0 else ''
    bbt = round(d['avg_bbls'] / d['avg_trucks'], 1) if d['avg_trucks'] > 0 else 0
    bbls_val = d['proj_bbls'] if d['days_remain'] > 0 else d['total_bbls']
    trucks_val = d['proj_trucks'] if d['days_remain'] > 0 else d['total_trucks']

    to = wb['Terminal Overview']
    to_row = find_month_row(to, d['month_abbr'], 21)
    safe_write(to, to_row, 1, f"{d['month_name']}{suffix}")
    safe_write(to, to_row, 2, 3)
    safe_write(to, to_row, 3, round(bbls_val))
    safe_write(to, to_row, 4, round(d['avg_bbls'], 0))
    safe_write(to, to_row, 5, round(trucks_val))
    safe_write(to, to_row, 6, d['avg_trucks'])
    safe_write(to, to_row, 7, bbt)
    safe_write(to, to_row, 8, d['pump_ute_combined'])
    safe_write(to, to_row, 9, d['rail_cap'])
    safe_write(to, to_row, 10, round(1 - d['rail_cap'], 3))

    om = wb['Operational Metrics']
    om_row = find_month_row(om, d['month_abbr'], 15)
    avail = d['days_actual'] * PUMP_AVAIL_HRS
    safe_write(om, om_row, 1, f"{d['month_name']}{suffix}")
    safe_write(om, om_row, 2, 3)
    safe_write(om, om_row, 3, d['p101_hrs'])
    safe_write(om, om_row, 4, round(d['p101_hrs'] / avail, 3) if avail > 0 else 0)
    safe_write(om, om_row, 5, d['p102_hrs'])
    safe_write(om, om_row, 6, round(d['p102_hrs'] / avail, 3) if avail > 0 else 0)
    safe_write(om, om_row, 7, d['p103_hrs'])
    safe_write(om, om_row, 8, round(d['p103_hrs'] / avail, 3) if avail > 0 else 0)
    safe_write(om, om_row, 9, d['total_pump_hrs'])
    safe_write(om, om_row, 10, d['pump_ute_combined'])

    ca = wb['Capacity Analysis']
    ca_row = find_month_row(ca, 'Current Rate', 29)
    safe_write(ca, ca_row, 1, f"Current Rate \u2014 {d['month_name']} Avg")
    safe_write(ca, ca_row, 6, round(d['avg_trucks'], 0))
    safe_write(ca, ca_row, 7, round(d['avg_bbls'], 0))
    safe_write(ca, ca_row, 9, round(bbls_val))

    wb.save(output_path)
    print(f"  External report saved: {os.path.basename(output_path)}")

# ════════════════════════════════════════════════════════════════════════════
# ATLAS SIGNALS — system health pulse + Tier 1 narrative (2026-04-13 revamp)
# ════════════════════════════════════════════════════════════════════════════

ATLAS_INTEGRITY_PATH = "/opt/nyx/data/integrity_check_results.json"

def fetch_atlas_pulse():
    """Read the latest integrity check (/opt/nyx/data/integrity_check_results.json).
    JSON schema: {timestamp, passed, failed, total, elapsed_seconds, checks: [{name, ok, detail}]}
    Returns dict with per-service up/down status + integrity (X/Y) + cadiz uptime."""
    pulse = {
        "integrity_passed": None, "integrity_total": None,
        "timestamp": None,
        "nyx_bot": "?", "nyx_bridge": "?", "memory_api": "?",
        "dashboard": "?", "training_portal": "?",
        "cadiz_pc": "?", "cadiz_uptime_h": None,
    }
    try:
        with open(ATLAS_INTEGRITY_PATH) as f:
            data = json.load(f)
        pulse["integrity_passed"] = data.get("passed")
        pulse["integrity_total"] = data.get("total")
        pulse["timestamp"] = (data.get("timestamp") or "")[:16]
        for c in data.get("checks", []):
            name = (c.get("name") or "").lower()
            ok = bool(c.get("ok"))
            detail = c.get("detail") or ""
            ico = "up" if ok else "down"
            if "nyx bot" in name:
                pulse["nyx_bot"] = ico
            elif "nyx bridge" in name:
                pulse["nyx_bridge"] = ico
            elif "memory api" in name:
                pulse["memory_api"] = ico
            elif "dashboard pwa" in name or "dashboard worker" in name:
                pulse["dashboard"] = ico
            elif "training portal" in name:
                pulse["training_portal"] = ico
            elif "cadiz pc" in name:
                pulse["cadiz_pc"] = ico
                m = re.search(r'uptime\s*([\d.]+)\s*h', detail, re.I)
                if m:
                    try:
                        pulse["cadiz_uptime_h"] = float(m.group(1))
                    except Exception:
                        pass
    except Exception as e:
        print(f"  ATLAS pulse fetch failed: {e}")
    return pulse


def _fetch_market_context():
    """Fetch current crude oil market headlines via Perplexity Sonar.
    Returns compact string or empty on any failure. Cost: ~$0.005/call, fires once/day in briefing.
    """
    try:
        import sys as _s; _s.path.insert(0, "/opt/nyx")
        from nyx_tools import execute_tool
        q = (
            "In 2-3 sentences, what moved the crude oil market in the past 24 hours? "
            "Focus on WTI/Brent price action, OPEC+ decisions, US inventory surprises, or "
            "geopolitical events relevant to a US midstream crude terminal. No bullet lists."
        )
        result = execute_tool("perplexity_research", {"query": q, "mode": "quick"})
        if not result or "error" in result.lower()[:20] or "not configured" in result.lower():
            return ""
        # Strip sources block — narrative prompt doesn't need inline URLs
        if "\n\nSources:" in result:
            result = result.split("\n\nSources:")[0]
        return result.strip()[:800]
    except Exception as _e:
        print(f"  ATLAS narrative: market context fetch failed ({_e})")
        return ""


def _push_atlas_to_dashboard(narrative, pulse, hero_kpis, market_context=""):
    """POST ATLAS data to Worker so VERA loader dashboard PWA can display it."""
    try:
        import urllib.request as _ur, json as _j, os as _os
        url = _os.environ.get("ATLAS_INGEST_URL", "https://cadiz-ops-worker.timiron-ops.workers.dev/api/atlas/ingest")
        token = _os.environ.get("ATLAS_INGEST_TOKEN", "")
        if not token:
            print("  ATLAS dashboard push: no ATLAS_INGEST_TOKEN, skipping")
            return
        payload = _j.dumps({
            "narrative": narrative or "",
            "pulse": pulse or None,
            "hero_kpis": hero_kpis or None,
            "market_context": market_context or "",
        }).encode()
        req = _ur.Request(url, data=payload, headers={
            "Content-Type": "application/json",
            "X-Atlas-Token": token,
            "User-Agent": "Timiron-Briefing/1.0",
        })
        r = _ur.urlopen(req, timeout=15)
        print(f"  ATLAS dashboard push: {r.status}")
    except Exception as _e:
        print(f"  ATLAS dashboard push failed: {_e}")

def generate_atlas_narrative(d, pu, yday_bbls, yday_trucks, combined_ute_pct):
    """Use Tier 1 (Max subscription via claude -p) to generate 2-3 sentence plain-English
    summary of yesterday's performance vs benchmarks. Zero cost. Returns empty on failure."""
    try:
        import sys as _sys
        _sys.path.insert(0, "/opt/nyx")
        from claude_max import query_max, MaxSubFailedError
    except Exception as e:
        print(f"  ATLAS narrative: import failed ({e})")
        return ""

    # Build compact facts block — include day-of-week explicitly so LLM doesn't guess
    mtd_vs_feb = (d['avg_bbls'] - FEB_2026_AVG) / FEB_2026_AVG * 100 if FEB_2026_AVG else 0
    yday_dt = d['yesterday_date']
    yday_label = yday_dt.strftime("%A, %B %d, %Y")  # "Sunday, April 12, 2026"
    facts = (
        f"Yesterday was {yday_label}.\n"
        f"- Volume: {yday_bbls:,.0f} BBLs across {yday_trucks} trucks\n"
        f"- Pump utilization (combined): {combined_ute_pct:.1f}% of 21-hr daily window\n"
        f"- Per-pump ute: P-101 {pu['P-101']['ute']}%  P-102 {pu['P-102']['ute']}%  P-103 {pu['P-103']['ute']}%\n"
        f"\nMonth-to-date ({d['days_actual']} days of {d['days_in_month']}):\n"
        f"- MTD BBLs: {d['total_bbls']:,.0f}\n"
        f"- Run rate: {d['avg_bbls']:,.0f} bbls/day ({mtd_vs_feb:+.1f}% vs Feb benchmark 11,200)\n"
        f"- Projected month-end: {d['proj_bbls']:,.0f} BBLs / {d['proj_trucks']:,.0f} trucks\n"
        f"- Rail capacity used: {d['rail_cap']*100:.1f}%\n"
    )
    # --- perplexity market context injection ---
    market_ctx = _fetch_market_context()
    market_block = f"\n\nCurrent crude market context (external):\n{market_ctx}\n" if market_ctx else ""
    if market_ctx:
        print(f"  Market context fetched ({len(market_ctx)} chars)")
    prompt = (
        "Write a 2-3 sentence plain-English briefing summary for Tyler, Director of Operations at Timiron "
        "Midstream Partners' Cadiz crude terminal. Tone: direct, operational, no filler. Lead with the most "
        "important fact from yesterday. If anything is notably above or below benchmark, call it out. "
        "If the market context below contains anything materially relevant to our ops (price swing, OPEC move, "
        "inventory surprise), weave one brief mention in. Avoid bullet lists — flowing sentences only. "
        "Do NOT repeat raw numbers the reader can see in the tables below; synthesize meaning.\n\nData:\n"
        + facts + market_block
    )
    try:
        sys_prompt = "You are Timiron's operations analyst. Terse, senior-level. No fluff."
        resp = query_max(
            user_text=prompt,
            system_prompt=sys_prompt,
            history=None,
            timeout=60,
        )
        return (resp or "").strip()
    except MaxSubFailedError as e:
        print(f"  ATLAS narrative: Tier 1 failed ({e}) -- skipping")
        return ""
    except Exception as e:
        print(f"  ATLAS narrative: error ({e}) -- skipping")
        return ""


# ════════════════════════════════════════════════════════════════════════════
# BUILD EMAIL HTML
# ════════════════════════════════════════════════════════════════════════════

def calc_switch_duration(start_str, end_str):
    try:
        fmt = '%I:%M%p'
        s = datetime.strptime(start_str.upper().replace(' ', ''), fmt)
        e = datetime.strptime(end_str.upper().replace(' ', ''), fmt)
        diff = (e - s).seconds // 60
        return f"{diff // 60}hr {diff % 60}min" if diff >= 60 else f"{diff}min"
    except:
        return ""

def build_cadiz_section(cadiz_data, carrier_actuals):
    """Build Cadiz Ops Activity + Carrier Performance HTML."""
    carrier_proj = cadiz_data.get('carrier_projections', {})
    maint_notes = cadiz_data.get('maintenance_notes', [])
    updates = cadiz_data.get('updates', [])

    has_content = any([updates, carrier_proj, carrier_actuals, maint_notes])
    if not has_content:
        return ""

    # Build update lines from raw emails
    update_html = ""
    for u in updates[:4]:
        body_short = u['body'][:120].replace('\n', ' ').strip()
        if body_short:
            subj = u['subject'].strip()
            update_html += f'<div class="kv"><span class="lbl">{subj[:30]}</span><span class="val">{body_short}</span></div>\n'

    # Maintenance
    maint_html = ""
    for note in maint_notes[:3]:
        maint_html += f'<div class="kv"><span class="lbl">Maintenance</span><span class="val" style="color:#90caf9;">{note[:150]}</span></div>\n'

    # Carrier table — known carriers first, then any new ones from actuals
    known_carriers = ['Badlands', 'KAG', 'Prop Logistics', 'BD Oil', '1st Choice Energy']
    extra_carriers = [c for c in carrier_actuals if c not in known_carriers]
    all_carriers = known_carriers + sorted(extra_carriers)
    total_proj = 0
    total_actual_trucks = 0
    total_actual_bbls = 0
    rows = ""
    for c in all_carriers:
        proj = carrier_proj.get(c, {})
        actual = carrier_actuals.get(c, {})
        pt = proj.get('trucks', 0)
        at = actual.get('trucks', 0)
        ab = actual.get('bbls', 0)
        total_proj += pt
        total_actual_trucks += at
        total_actual_bbls += ab

        proj_str = str(pt) if pt > 0 else '<span style="color:#666;font-style:italic;">\u2014</span>'
        actual_str = str(at) if at > 0 else '<span style="color:#666;">0</span>'
        bbls_str = f"{ab:,.0f}" if at > 0 else '<span style="color:#666;">\u2014</span>'

        if pt > 0 and at > 0:
            var = at - pt
            var_str = f'<span style="color:{"#4caf50" if var >= 0 else "#ef5350"}">{var:+d}</span>'
        else:
            var_str = '<span style="color:#666;">\u2014</span>'

        rows += f'<tr><td>{c}</td><td>{proj_str}</td><td>{actual_str}</td><td>{bbls_str}</td><td>{var_str}</td></tr>\n'

    dash = "\u2014"
    proj_display = total_proj if total_proj else dash
    total_row = f'<tr class="tot"><td>Total</td><td>{proj_display}</td><td>{total_actual_trucks}</td><td>{total_actual_bbls:,.0f} BBLs</td><td>{dash}</td></tr>'

    return f"""
<div class="section">
  <div class="sec-head">\U0001f4e1 Cadiz Ops Activity</div>
  {update_html}
  {maint_html}
  <div class="sec-head" style="margin-top:10px;">\U0001f69b Carrier Performance</div>
  <table>
    <tr><th>Carrier</th><th>Projected</th><th>Actual</th><th>Actual BBLs</th><th>Variance</th></tr>
    {rows}
    {total_row}
  </table>
</div>"""

# ════════════════════════════════════════════════════════════════════════════
# QUICKBOOKS TIME - CREW HOURS
# ════════════════════════════════════════════════════════════════════════════

QBT_API = "https://rest.tsheets.com/api/v1"

QBT_DAY_CREW = [
    "Cameron Betz", "Shawn Osborn Jr.", "Shane Young", "William Glover",
    "Austin Tredway", "Gregory Bates", "Jared Wright", "Shawn Osborn Sr.",
]
QBT_NIGHT_CREW = [
    "Jonathan Williams", "Daniel Hough", "Bryan Deoss", "Dustin Fletcher",
    "Jacob Diloreto", "Nathaniel Medel", "Christopher Wright",
]
QBT_ROSTER = QBT_DAY_CREW + QBT_NIGHT_CREW
QBT_SHAWN_MAP = {
    "gosborn20@gmail.com": "Shawn Osborn Jr.",
    "osbornshawn25@gmail.com": "Shawn Osborn Sr.",
}

def qbt_api_get(endpoint, params):
    """GET with pagination from QBT API."""
    all_results = {}
    page = 1
    while True:
        p = dict(params)
        p["page"] = page
        p["per_page"] = 200
        r = requests.get(f"{QBT_API}/{endpoint}",
                         headers={"Authorization": f"Bearer {QBT_TOKEN}"},
                         params=p, timeout=30)
        r.raise_for_status()
        data = r.json()
        results = data.get("results", {}).get(endpoint, {})
        if not results:
            break
        all_results.update(results)
        if not data.get("more", False):
            break
        page += 1
    return all_results

def fetch_qbt_crew_hours():
    """Fetch WTD hours from QuickBooks Time. Returns (rows, week_label) or (None, None) on failure."""
    if not QBT_TOKEN:
        print("  QBT_TOKEN not set, skipping crew hours.")
        return None, None

    try:
        # Get users
        raw_users = qbt_api_get("users", {"active": "yes"})
        users = {}
        for uid, u in raw_users.items():
            first = (u.get("first_name") or "").strip()
            last = (u.get("last_name") or "").strip()
            email = (u.get("email") or "").strip().lower()
            full = f"{first} {last}".strip()
            if email in QBT_SHAWN_MAP:
                full = QBT_SHAWN_MAP[email]
            users[str(uid)] = full

        # Current week Mon-Sun
        from zoneinfo import ZoneInfo
        ET = ZoneInfo("America/New_York")
        now_et = datetime.now(ET)
        today_et = now_et.date()
        monday = today_et - timedelta(days=today_et.weekday())
        sunday = monday + timedelta(days=6)

        # Fetch completed timesheets
        completed = qbt_api_get("timesheets", {
            "start_date": monday.isoformat(),
            "end_date": sunday.isoformat(),
        })

        # Fetch active shifts (check yesterday too for overnight)
        yesterday = monday - timedelta(days=1)
        active_y = qbt_api_get("timesheets", {"on_the_clock": "yes", "start_date": yesterday.isoformat()})
        active_t = qbt_api_get("timesheets", {"on_the_clock": "yes", "start_date": monday.isoformat()})
        for tid, ts in {**active_y, **active_t}.items():
            completed[tid] = ts

        # Aggregate per employee
        emp = {}
        for tid, ts in completed.items():
            uid = str(ts.get("user_id", ""))
            name = users.get(uid, f"Unknown ({uid})")
            duration = ts.get("duration", 0)
            if duration == 0 and ts.get("end", "") == "":
                start_str = ts.get("start", "")
                if start_str:
                    start_utc = datetime.fromisoformat(start_str.replace("Z", "+00:00"))
                    duration = int((datetime.now(timezone.utc) - start_utc).total_seconds())
            if name not in emp:
                emp[name] = 0
            emp[name] += duration

        # Build rows sorted by roster
        rows = []
        for name in QBT_ROSTER:
            if name in emp:
                total_hrs = round(emp[name] / 3600, 1)
                reg = min(total_hrs, 40)
                ot = round(max(total_hrs - 40, 0), 1)
                shift = "Day" if name in QBT_DAY_CREW else "Night"
                rows.append({"name": name, "shift": shift, "total": total_hrs, "reg": reg, "ot": ot})
        # Add anyone not in roster
        for name, secs in emp.items():
            if name not in QBT_ROSTER:
                total_hrs = round(secs / 3600, 1)
                rows.append({"name": name, "shift": "-", "total": total_hrs,
                             "reg": min(total_hrs, 40), "ot": round(max(total_hrs - 40, 0), 1)})

        week_label = f"{fmt_date_short(monday)}-{fmt_date_short(today_et)}"
        return rows, week_label

    except Exception as e:
        print(f"  QBT fetch failed: {e}")
        return None, None

def build_crew_hours_html(rows, week_label):
    """Build dark-themed HTML section for crew WTD hours."""
    if not rows:
        return ""

    total_hrs = sum(r["total"] for r in rows)
    total_ot = sum(r["ot"] for r in rows)

    table_rows = ""
    for r in rows:
        ot_style = ""
        if r["ot"] > 25:
            ot_style = ' style="color:#ef5350;font-weight:bold"'
        elif r["ot"] > 0:
            ot_style = ' style="color:#FFD100"'
        flag = ' <span style="color:#ef5350">!!</span>' if r["total"] > 60 else ""
        table_rows += (
            f'<tr><td>{r["name"]}</td><td>{r["shift"]}</td>'
            f'<td>{r["total"]:.1f}{flag}</td><td>{r["reg"]:.1f}</td>'
            f'<td{ot_style}>{r["ot"]:.1f}</td></tr>\n'
        )

    day_rows = [r for r in rows if r["shift"] == "Day"]
    night_rows = [r for r in rows if r["shift"] == "Night"]
    day_avg = sum(r["total"] for r in day_rows) / len(day_rows) if day_rows else 0
    night_avg = sum(r["total"] for r in night_rows) / len(night_rows) if night_rows else 0

    ot_note = ""
    if len(night_rows) < 8 and night_avg > day_avg:
        ot_note = (
            f'<div class="flag" style="margin-top:8px;">Night shift ({len(night_rows)}) averaging '
            f'{night_avg:.1f} hrs vs Day ({len(day_rows)}) at {day_avg:.1f} hrs. '
            f'Target schedule is 5-on/3-off (8 per shift). Night running short-handed '
            f'— OT normalizes once night reaches 8.</div>'
        )

    return f"""<div class="section">
  <div class="sec-head">Crew Hours WTD - {week_label}</div>
  <table>
    <tr><th>Name</th><th>Shift</th><th>Total</th><th>Reg</th><th>OT</th></tr>
    {table_rows}
    <tr class="tot"><td>TOTALS ({len(rows)})</td><td></td><td>{total_hrs:.1f}</td><td>{sum(r['reg'] for r in rows):.1f}</td><td>{total_ot:.1f}</td></tr>
  </table>
  {ot_note}
  <div style="color:#555;font-size:10px;margin-top:6px;">Source: QuickBooks Time API (includes active shifts)</div>
</div>"""

# ════════════════════════════════════════════════════════════════════════════
# HERA — Commercial Intelligence Section
# ════════════════════════════════════════════════════════════════════════════

def build_hera_section():
    """Fetch HERA commercial intel and render as HTML section for the briefing.
    Returns empty string if HERA API is unreachable (graceful degradation)."""
    import requests as _hera_requests
    HERA_BASE = "http://127.0.0.1:5697"
    try:
        # Fetch last hunt stats
        hunt_resp = _hera_requests.get(f"{HERA_BASE}/hunt/last", timeout=5)
        hunt_resp.raise_for_status()
        hunt_data = hunt_resp.json()
        hunt = hunt_data.get("last_hunt", hunt_data)  # unwrap envelope
        hunt_ts = hunt.get("ts", hunt.get("timestamp", "unknown"))
        hunt_scanned = hunt.get("items_scanned", 0)
        hunt_matched = hunt.get("items_matched", 0)

        # Fetch new opportunities
        opps_resp = _hera_requests.get(
            f"{HERA_BASE}/opportunities",
            params={"status": "new", "limit": "10"},
            timeout=5,
        )
        opps_resp.raise_for_status()
        opps_data = opps_resp.json()
        opps = opps_data.get("opportunities", opps_data) if isinstance(opps_data, dict) else opps_data

        # Sort by urgency: HIGH > MEDIUM > LOW
        urgency_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
        opps = sorted(opps, key=lambda o: urgency_order.get(o.get("urgency", "LOW").upper(), 3))

        # Build opportunity rows
        if opps:
            opp_rows = ""
            for o in opps:
                urg = o.get("urgency", "LOW").upper()
                urg_color = {"HIGH": "#ef5350", "MEDIUM": "#ffb74d", "LOW": "#4caf50"}.get(urg, "#888")
                otype = o.get("opportunity_type", "General")
                liner = o.get("one_liner", "No description")
                action = o.get("suggested_action", "Review")
                opp_rows += (
                    f'<div style="margin-bottom:10px;padding:8px;background:#1a1a1a;border-radius:4px;">'
                    f'<div><span style="color:{urg_color};font-weight:bold;">[{urg}]</span> '
                    f'<span style="color:#90caf9;">{otype}</span>: {liner}</div>'
                    f'<div style="color:#888;font-size:11px;margin-top:3px;">Action: {action}</div>'
                    f'</div>'
                )
            opp_html = opp_rows
            count_label = f"{len(opps)} new opportunit{'y' if len(opps) == 1 else 'ies'}"
        else:
            opp_html = '<div style="color:#888;font-style:italic;">No new commercial opportunities since last briefing.</div>'
            count_label = "0 new"

        return (
            '<div style="margin:24px 0;padding:14px;background:#0d1117;border:1px solid #30363d;'
            'border-radius:8px;font-family:Segoe UI,Arial,sans-serif;color:#c9d1d9;">'
            '<div style="color:#58a6ff;font-weight:bold;font-size:14px;margin-bottom:8px;">'
            'HERA &mdash; Commercial Intelligence</div>'
            f'<div style="color:#888;font-size:11px;margin-bottom:12px;">'
            f'Last hunt: {hunt_ts} | {hunt_scanned} scanned | {hunt_matched} matches | {count_label}</div>'
            f'{opp_html}'
            '</div>'
        )
    except Exception as e:
        print(f"  HERA section skipped (API unreachable): {e}")
        return ""


def build_rail_section():
    """Fetch W&LE rail pipeline data and render as HTML section.
    Returns empty string if HERA API is unreachable (graceful degradation)."""
    import requests as _rail_requests
    HERA_BASE = "http://127.0.0.1:5697"
    try:
        resp = _rail_requests.get(f"{HERA_BASE}/rail/summary", timeout=5)
        resp.raise_for_status()
        data = resp.json()

        # Cars by status
        status_counts = data.get("by_status", data.get("cars_by_status", {}))
        status_colors = {"enroute": "#42a5f5", "yard": "#ffb74d", "industry": "#66bb6a"}
        status_pills = ""
        for st in ["enroute", "yard", "industry"]:
            cnt = status_counts.get(st, 0)
            color = status_colors.get(st, "#888")
            status_pills += (
                f'<span style="display:inline-block;margin-right:12px;padding:4px 10px;'
                f'background:{color}22;border:1px solid {color};border-radius:4px;'
                f'color:{color};font-weight:bold;font-size:13px;">'
                f'{st.upper()}: {cnt}</span>'
            )

        # Pipeline positions
        pipeline = data.get("pipeline", data.get("positions", {}))
        stops = ["BREWSTER", "SHANNONRN", "MINGOJCT", "NELMS"]
        pipe_html = '<div style="margin-top:10px;display:flex;align-items:center;gap:0;">'
        for i, stop in enumerate(stops):
            cnt = pipeline.get(stop, pipeline.get(stop.lower(), 0))
            pipe_html += (
                f'<div style="text-align:center;padding:6px 12px;background:#1a1a1a;'
                f'border-radius:4px;min-width:80px;">'
                f'<div style="color:#90caf9;font-size:11px;">{stop}</div>'
                f'<div style="color:#fff;font-weight:bold;font-size:16px;">{cnt}</div>'
                f'</div>'
            )
            if i < len(stops) - 1:
                pipe_html += '<div style="color:#555;font-size:16px;padding:0 4px;">&rarr;</div>'
        pipe_html += "</div>"

        report_date = data.get("report_date", data.get("date", "unknown"))

        return (
            '<div style="margin:24px 0;padding:14px;background:#0d1117;border:1px solid #30363d;'
            'border-radius:8px;font-family:Segoe UI,Arial,sans-serif;color:#c9d1d9;">'
            '<div style="color:#58a6ff;font-weight:bold;font-size:14px;margin-bottom:8px;">'
            'Rail Pipeline &mdash; W&amp;LE Tracer</div>'
            f'<div style="color:#888;font-size:11px;margin-bottom:10px;">Report date: {report_date}</div>'
            f'<div style="margin-bottom:10px;">{status_pills}</div>'
            f'{pipe_html}'
            '</div>'
        )
    except Exception as e:
        print(f"  Rail Pipeline section skipped (API unreachable): {e}")
        return ""


def build_invoice_section():
    """Fetch invoice/spend summary and render as HTML section.
    Returns empty string if HERA API is unreachable (graceful degradation)."""
    import requests as _inv_requests
    HERA_BASE = "http://127.0.0.1:5697"
    try:
        resp = _inv_requests.get(f"{HERA_BASE}/invoices/summary", timeout=5)
        resp.raise_for_status()
        data = resp.json()

        total_invoices = data.get("total_invoices", data.get("count", 0))
        total_spend = data.get("total_spend", data.get("total_extracted_spend", 0))
        date_range = data.get("date_range", data.get("range", "unknown"))
        if isinstance(date_range, dict):
            date_range = f"{date_range.get('start', '?')} to {date_range.get('end', '?')}"
        top_vendors = data.get("top_vendors", data.get("vendors_by_spend", []))[:5]

        # Format spend
        if isinstance(total_spend, (int, float)):
            spend_fmt = f"${total_spend:,.2f}"
        else:
            spend_fmt = str(total_spend)

        vendor_rows = ""
        for v in top_vendors:
            vname = v.get("vendor", v.get("name", "Unknown"))
            vamt = v.get("spend", v.get("amount", v.get("total", 0)))
            if isinstance(vamt, (int, float)):
                vamt_fmt = f"${vamt:,.2f}"
            else:
                vamt_fmt = str(vamt)
            vendor_rows += (
                f'<div style="display:flex;justify-content:space-between;padding:5px 8px;'
                f'background:#1a1a1a;border-radius:4px;margin-bottom:4px;">'
                f'<span style="color:#c9d1d9;">{vname}</span>'
                f'<span style="color:#66bb6a;font-weight:bold;">{vamt_fmt}</span>'
                f'</div>'
            )

        if not vendor_rows:
            vendor_rows = '<div style="color:#888;font-style:italic;">No vendor data available.</div>'

        return (
            '<div style="margin:24px 0;padding:14px;background:#0d1117;border:1px solid #30363d;'
            'border-radius:8px;font-family:Segoe UI,Arial,sans-serif;color:#c9d1d9;">'
            '<div style="color:#58a6ff;font-weight:bold;font-size:14px;margin-bottom:8px;">'
            "Financial Intelligence &mdash; Kelly's Invoices</div>"
            f'<div style="color:#888;font-size:11px;margin-bottom:10px;">'
            f'{total_invoices} invoices | {spend_fmt} total spend | {date_range}</div>'
            f'<div style="font-size:12px;color:#888;margin-bottom:6px;font-weight:bold;">Top 5 Vendors by Spend</div>'
            f'{vendor_rows}'
            '</div>'
        )
    except Exception as e:
        print(f"  Invoice section skipped (API unreachable): {e}")
        return ""


def build_email_html(d, dash_name, ext_name, cadiz_section="", crew_section="", atlas_pulse=None, atlas_narrative=""):
    """Revamped 2026-04-13 — hero KPI cards, ATLAS signals (pulse + narrative), sparkline.
    Keeps all existing data; reshapes top of email for mobile-first, glance-able digest."""
    pu = d['pump_ute']
    dt_str = fmt_date(d['yesterday_date'])
    today_str = datetime.now().strftime('%A, %B %d, %Y')  # cross-platform

    yday_bbls = sum(v['bbls'] for v in pu.values())
    yday_trucks = sum(v['loads'] for v in pu.values())
    yday_splits = sum(v['splits'] for v in pu.values())
    yday_rt = sum(v['runtime'] for v in pu.values())
    combined_ute = yday_rt / (PUMP_AVAIL_HRS * 3) * 100 if yday_rt > 0 else 0
    bbl_hr_comb = yday_bbls / yday_rt if yday_rt > 0 else 0
    bbl_per_truck = yday_bbls / yday_trucks if yday_trucks > 0 else 0

    vs_run = (yday_bbls - d['avg_bbls']) / d['avg_bbls'] * 100 if d['avg_bbls'] > 0 else 0
    vs_feb = d['proj_bbls'] - FEB_2026_TOTAL
    mtd_vs = (d['avg_bbls'] - FEB_2026_AVG) / FEB_2026_AVG * 100
    yday_rev = rev_per_day(yday_bbls)

    avg_api = d.get('avg_api_gravity', 0)
    avg_bsw = d.get('avg_bsw', 0)
    mabbr = d['month_abbr']
    month_label = d['month_name']

    def arrow(val):
        """Trend arrow with color for a % delta."""
        if val >= 1:
            return f'<span style="color:#4caf50">&#9650; {val:+.1f}%</span>'
        if val <= -1:
            return f'<span style="color:#ef5350">&#9660; {val:+.1f}%</span>'
        return f'<span style="color:#888">&#9632; {val:+.1f}%</span>'

    def badge(val):
        c = '#4caf50' if val >= 0 else '#ef5350'
        return f'<span style="color:{c}">{val:+.1f}%</span>'

    def badge_abs(val):
        c = '#4caf50' if val >= 0 else '#ef5350'
        return f'<span style="color:{c}">{val:+,.0f} BBLs</span>'

    def hero_card(label, value, sub=""):
        return (
            '<td style="width:25%;padding:8px;text-align:center;vertical-align:top;">'
            f'<div style="color:#888;font-size:10px;text-transform:uppercase;letter-spacing:1px;">{label}</div>'
            f'<div style="color:#fff;font-size:22px;font-weight:bold;margin:4px 0;">{value}</div>'
            f'<div style="color:#777;font-size:10px;">{sub}</div>'
            '</td>'
        )

    # 5-Day Trend
    daily_data = d.get('daily_data', [])
    last_5 = daily_data[-5:] if len(daily_data) >= 5 else daily_data
    trend_rows = ""
    for i, dd in enumerate(last_5):
        bpt = dd['bbls'] / dd['trucks'] if dd['trucks'] > 0 else 0
        vs = (dd['bbls'] - d['avg_bbls']) / d['avg_bbls'] * 100 if d['avg_bbls'] > 0 else 0
        vs_col = '#4caf50' if vs >= 0 else '#ef5350'
        if i > 0:
            prev = last_5[i - 1]['bbls']
            dod = (dd['bbls'] - prev) / prev * 100 if prev > 0 else 0
            dod_col = '#4caf50' if dod >= 0 else '#ef5350'
            dod_str = f'<span style="color:{dod_col}">{dod:+.1f}%</span>'
        else:
            dod_str = '<span style="color:#666">\u2014</span>'
        day_label = f"{dd['date'].strftime('%b')} {dd['date'].day} ({dd['day_name']})"
        trend_rows += f'<tr><td>{day_label}</td><td>{dd["bbls"]:,.0f}</td><td>{dd["trucks"]}</td><td>{bpt:.1f}</td><td style="color:{vs_col}">{vs:+.1f}%</td><td>{dod_str}</td></tr>\n'

    wd = [dd for dd in last_5 if dd['day_name'] not in ('Sat', 'Sun')]
    we = [dd for dd in last_5 if dd['day_name'] in ('Sat', 'Sun')]
    wd_avg = sum(dd['bbls'] for dd in wd) / len(wd) if wd else 0
    we_avg = sum(dd['bbls'] for dd in we) / len(we) if we else 0
    trend_note = f"weekday: {wd_avg:,.0f}/day ({len(wd)})" if wd else ""
    if we:
        trend_note += f" | weekend: {we_avg:,.0f}/day ({len(we)})"

    # Weekly Breakdown
    weekly_data = d.get('weekly_data', [])
    weekly_rows = ""
    for w in weekly_data:
        w_vs = (w['avg_bbls'] - d['avg_bbls']) / d['avg_bbls'] * 100 if d['avg_bbls'] > 0 else 0
        w_col = '#4caf50' if w_vs >= 0 else '#ef5350'
        weekly_rows += f'<tr><td>{w["label"]}</td><td>{w["bbls"]:,.0f}</td><td>{w["trucks"]}</td><td>{w["days"]}</td><td>{w["avg_bbls"]:,.1f}</td><td>{w["avg_bpt"]:.1f}</td><td style="color:{w_col}">{w_vs:+.1f}%</td></tr>\n'

    # Flags
    flags = []
    utes = {k: v['ute'] for k, v in pu.items()}
    max_p = max(utes, key=utes.get)
    min_p = min(utes, key=utes.get)
    if utes[max_p] - utes[min_p] > 10:
        flags.append(('yellow', f"{min_p} ute {utes[min_p]}% vs {max_p} {utes[max_p]}% \u2014 load imbalance across pumps."))
    low_hr = [f"{k} at {v['bbl_hr']:.0f}" for k, v in pu.items() if 0 < v['bbl_hr'] < 430]
    if low_hr:
        flags.append(('yellow', f"BBLs/hr below Feb avg (438): {', '.join(low_hr)}."))
    split_pct = (yday_splits / yday_trucks * 100) if yday_trucks > 0 else 0
    if split_pct <= 27:
        flags.append(('grn', f"Split count {yday_splits}/{yday_trucks} ({split_pct:.1f}%) \u2014 near historical avg (24%)."))

    flags_html = ""
    for ftype, msg in flags:
        cls = 'flag red' if ftype == 'red' else ('flag grn' if ftype == 'grn' else 'flag')
        flags_html += f'<div class="{cls}">{msg}</div>\n'

    forecast_label = f"{month_label} Forecast ({d['days_actual']}-Day Rate)" if d['days_remain'] > 0 else f"{month_label} Final ({d['days_actual']}-Day Actuals)"

    # === ATLAS pulse strip ===
    ap = atlas_pulse or {}
    def pulse_dot(status):
        if status == "up":
            return '<span style="color:#4caf50;font-weight:bold;">&#9679;</span>'
        if status == "down":
            return '<span style="color:#ef5350;font-weight:bold;">&#9679;</span>'
        return '<span style="color:#666;">&#9679;</span>'
    ap_nyx = pulse_dot(ap.get("nyx_bot", "?"))
    ap_br = pulse_dot(ap.get("nyx_bridge", "?"))
    ap_mem = pulse_dot(ap.get("memory_api", "?"))
    ap_dash = pulse_dot(ap.get("dashboard", "?"))
    ap_train = pulse_dot(ap.get("training_portal", "?"))
    ap_cadiz = pulse_dot(ap.get("cadiz_pc", "?"))
    ap_cadiz_up = f'{ap.get("cadiz_uptime_h", 0):.0f}h' if ap.get("cadiz_uptime_h") else "?"
    ap_int_p = ap.get("integrity_passed")
    ap_int_t = ap.get("integrity_total")
    if ap_int_p is not None and ap_int_t:
        ap_int_txt = f"{ap_int_p}/{ap_int_t}"
        ap_int_col = "#4caf50" if ap_int_p == ap_int_t else "#ef5350"
    else:
        ap_int_txt = "?"
        ap_int_col = "#666"
    ap_time = ap.get("timestamp") or ""

    narrative_html = ""
    if atlas_narrative:
        narrative_html = (
            '<div class="narr">'
            f'<div class="narr-label">ATLAS BRIEFING</div>'
            f'<div class="narr-body">{atlas_narrative}</div>'
            '</div>'
        )

    pulse_html = (
        '<div class="pulse">'
        f'<span class="pulse-label">ATLAS PULSE</span> &nbsp; '
        f'{ap_nyx} NYX &nbsp; {ap_br} Bridge &nbsp; {ap_mem} Mem &nbsp; '
        f'{ap_dash} Dash &nbsp; {ap_train} Train &nbsp; {ap_cadiz} Cadiz PC ({ap_cadiz_up}) &nbsp; '
        f'<span style="color:{ap_int_col};">Integrity {ap_int_txt}</span>'
        f' &nbsp;<span class="pulse-time">{ap_time}</span>'
        '</div>'
    )

    # Hero 4-card grid — BBLs, Trucks, Util%, Est Revenue
    hero_html = (
        '<table class="hero" width="100%" cellspacing="0" cellpadding="0"><tr>'
        + hero_card("Yesterday BBLs", f"{yday_bbls:,.0f}", f"{arrow(vs_run)} vs run rate")
        + hero_card("Trucks", f"{yday_trucks}", f"{bbl_per_truck:.0f} BBL / truck")
        + hero_card("Pump Util", f"{combined_ute:.0f}%", f"{yday_rt:.1f} run-hrs")
        + hero_card("Est Revenue", f"${yday_rev/1000:,.1f}k", f"{bbl_hr_comb:.0f} BBL / hr")
        + '</tr></table>'
    )

    # Pre-compose flags block (moved up in layout)
    flags_block = (
        '<div class="section">'
        '<div class="sec-head">Flags &amp; Attention</div>'
        f'{flags_html}'
        '<div class="flag grn">&#128206; Both Excel files attached (internal + external report).</div>'
        '</div>'
    ) if flags_html else (
        '<div class="section">'
        '<div class="sec-head">Flags &amp; Attention</div>'
        '<div class="flag grn">No operational flags raised. &#128206; Both Excel files attached.</div>'
        '</div>'
    )

    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  body{{background:#1a1a1a;font-family:'Courier New',monospace;color:#e0e0e0;margin:0;padding:20px;}}
  .wrap{{max-width:680px;margin:0 auto;}}
  .title{{color:#fff;font-size:16px;font-weight:bold;border-bottom:1px solid #444;padding-bottom:8px;margin-bottom:6px;}}
  .sub{{color:#666;font-size:11px;margin-bottom:14px;}}
  .narr{{background:#1e2a1e;border-left:3px solid #4caf50;padding:10px 14px;margin-bottom:12px;border-radius:2px;}}
  .narr-label{{color:#4caf50;font-size:9px;text-transform:uppercase;letter-spacing:1.2px;margin-bottom:4px;font-weight:bold;}}
  .narr-body{{color:#e0e0e0;font-size:13px;line-height:1.55;font-family:-apple-system,BlinkMacSystemFont,Segoe UI,sans-serif;}}
  .pulse{{background:#1f1f1f;border:1px solid #333;padding:8px 12px;margin-bottom:12px;border-radius:2px;font-size:11px;color:#ccc;}}
  .pulse-label{{color:#777;font-weight:bold;letter-spacing:1px;}}
  .pulse-time{{color:#555;font-size:9px;}}
  table.hero{{border-collapse:collapse;margin-bottom:12px;background:#242424;border:1px solid #333;border-radius:3px;}}
  table.hero td{{border-right:1px solid #2e2e2e;}}
  table.hero tr td:last-child{{border-right:none;}}
  .section{{background:#242424;border:1px solid #333;border-radius:3px;padding:13px 15px;margin-bottom:12px;}}
  .sec-head{{color:#999;font-size:10px;text-transform:uppercase;letter-spacing:1px;margin-bottom:9px;border-bottom:1px solid #2e2e2e;padding-bottom:5px;}}
  .kv{{display:flex;justify-content:space-between;padding:3px 0;font-size:13px;border-bottom:1px solid #2a2a2a;}}
  .kv:last-child{{border-bottom:none;}}
  .lbl{{color:#888;}}.val{{color:#ddd;font-weight:bold;}}
  table{{width:100%;border-collapse:collapse;font-size:12px;margin-top:2px;}}
  th{{color:#777;font-weight:normal;text-align:left;padding:4px 6px;border-bottom:1px solid #333;font-size:10px;}}
  td{{padding:5px 6px;border-bottom:1px solid #2a2a2a;color:#ccc;}}
  tr.tot td{{color:#fff;font-weight:bold;background:#2d2d2d;border-top:1px solid #444;}}
  .flag{{padding:6px 10px;margin-bottom:6px;border-left:3px solid #FFD100;background:#2a2700;font-size:12px;border-radius:2px;}}
  .flag.grn{{border-left-color:#4caf50;background:#162016;}}
  .flag.red{{border-left-color:#ef5350;background:#2a1616;}}
  .foot{{color:#444;font-size:10px;text-align:center;margin-top:14px;border-top:1px solid #2a2a2a;padding-top:10px;}}
</style>
</head><body><div class="wrap">
<div class="title">\U0001f4ca Timiron Daily Briefing &nbsp;|&nbsp; Cadiz Terminal</div>
<div class="sub">{today_str} &nbsp;\u00b7&nbsp; Reporting {dt_str}</div>

{narrative_html}

{hero_html}

{pulse_html}

{flags_block}

{cadiz_section}

<div class="section">
  <div class="sec-head">Pump Detail \u2014 Yesterday ({dt_str})</div>
  <table>
    <tr><th>Pump</th><th>Loads</th><th>Splits</th><th>Hrs</th><th>Ute%</th><th>BBLs</th><th>BBL/Hr</th></tr>
    <tr><td>P-101</td><td>{pu['P-101']['loads']}</td><td>{pu['P-101']['splits']}</td><td>{pu['P-101']['runtime']}</td><td>{pu['P-101']['ute']}%</td><td>{pu['P-101']['bbls']:,.0f}</td><td>{pu['P-101']['bbl_hr']:.0f}</td></tr>
    <tr><td>P-102</td><td>{pu['P-102']['loads']}</td><td>{pu['P-102']['splits']}</td><td>{pu['P-102']['runtime']}</td><td>{pu['P-102']['ute']}%</td><td>{pu['P-102']['bbls']:,.0f}</td><td>{pu['P-102']['bbl_hr']:.0f}</td></tr>
    <tr><td>P-103</td><td>{pu['P-103']['loads']}</td><td>{pu['P-103']['splits']}</td><td>{pu['P-103']['runtime']}</td><td>{pu['P-103']['ute']}%</td><td>{pu['P-103']['bbls']:,.0f}</td><td>{pu['P-103']['bbl_hr']:.0f}</td></tr>
    <tr class="tot"><td>Combined</td><td>{yday_trucks}</td><td>{yday_splits}</td><td>{yday_rt:.2f}</td><td>{combined_ute:.1f}%</td><td>{yday_bbls:,.0f}</td><td>{bbl_hr_comb:.0f}</td></tr>
  </table>
  <div style="color:#555;font-size:10px;margin-top:6px;">Yesterday per-pump Ute% = runtime / 21 daily-avail hrs (24hr minus 3hr rail switch). MTD pump-util in the xlsx uses calendar-hours to stay comparable to prior months.</div>
</div>

<div class="section">
  <div class="sec-head">Crude Quality \u2014 Yesterday</div>
  <div class="kv"><span class="lbl">Avg API Gravity</span><span class="val">{avg_api:.2f}</span></div>
  <div class="kv"><span class="lbl">Avg BSW</span><span class="val">{avg_bsw:.3f}</span></div>
</div>

<div class="section">
  <div class="sec-head">5-Day Trend</div>
  <table>
    <tr><th>Date</th><th>BBLs</th><th>Trk</th><th>B/T</th><th>vs Avg</th><th>DoD</th></tr>
    {trend_rows}
  </table>
  <div style="color:#555;font-size:10px;margin-top:6px;">{trend_note}</div>
</div>

<div class="section">
  <div class="sec-head">Month-to-Date ({mabbr} 1\u2013{d['yesterday_date'].day})</div>
  <div class="kv"><span class="lbl">MTD Actuals</span><span class="val">{d['total_bbls']:,.0f} BBLs</span></div>
  <div class="kv"><span class="lbl">Daily Avg ({d['days_actual']} days)</span><span class="val">{d['avg_bbls']:,.1f} bbls/day</span></div>
  <div class="kv"><span class="lbl">vs Feb (11,200/day)</span><span class="val">{badge(mtd_vs)}</span></div>
  <div class="kv"><span class="lbl">Rail Cap Used</span><span class="val" style="color:#90caf9">{d['rail_cap']*100:.1f}%</span></div>
  <div class="kv"><span class="lbl">Days Remaining</span><span class="val">{d['days_remain']}</span></div>
</div>

<div class="section">
  <div class="sec-head">Weekly Breakdown</div>
  <table>
    <tr><th>Week</th><th>BBLs</th><th>Trk</th><th>Days</th><th>Avg</th><th>B/T</th><th>vs Avg</th></tr>
    {weekly_rows}
  </table>
</div>

<div class="section">
  <div class="sec-head">{forecast_label}</div>
  <div class="kv"><span class="lbl">Run Rate</span><span class="val">{d['avg_bbls']:,.1f} bbls/day</span></div>
  <div class="kv"><span class="lbl">{'Projected' if d['days_remain'] > 0 else 'Total'} BBLs</span><span class="val">{d['proj_bbls']:,.0f}</span></div>
  <div class="kv"><span class="lbl">{'Projected' if d['days_remain'] > 0 else 'Total'} Trucks</span><span class="val">{'~' if d['days_remain'] > 0 else ''}{d['proj_trucks']:,.0f}</span></div>
  <div class="kv"><span class="lbl">vs Feb (313,600 BBLs)</span><span class="val">{badge_abs(vs_feb)}</span></div>
</div>

{crew_section}

</div>
<div class="foot">
  Timiron Midstream Partners \u00b7 Cadiz Terminal, OH \u00b7 ATLAS briefing\u2014v5<br>
  {dash_name} \u00b7 {ext_name}
</div>
</div></body></html>"""

# ════════════════════════════════════════════════════════════════════════════
# SEND EMAIL
# ════════════════════════════════════════════════════════════════════════════

def send_via_gmail(subject, html_body, attachment_paths):
    msg = MIMEMultipart('mixed')
    msg['From'] = GMAIL_ADDRESS
    msg['To'] = ', '.join(RECIPIENTS)
    msg['Subject'] = subject
    msg.attach(MIMEText(html_body, 'html'))

    for path in attachment_paths:
        with open(path, 'rb') as f:
            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(path))
        msg.attach(part)

    with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as server:
        server.starttls()
        server.login(GMAIL_ADDRESS, GMAIL_APP_PASS)
        server.sendmail(GMAIL_ADDRESS, RECIPIENTS, msg.as_string())

    print(f"  Email sent to {', '.join(RECIPIENTS)} with {len(attachment_paths)} attachments.")

def send_error_email(error_msg):
    """Send a brief error notification if the briefing fails."""
    try:
        msg = MIMEMultipart()
        msg['From'] = GMAIL_ADDRESS
        msg['To'] = RECIPIENTS[0]  # Just Tyler
        msg['Subject'] = f"\u26a0 Timiron Briefing FAILED | {date.today().strftime('%b %d')}"
        body = f"The daily briefing pipeline failed.\n\n{error_msg}\n\nCheck GitHub Actions for details."
        msg.attach(MIMEText(body, 'plain'))
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as server:
            server.starttls()
            server.login(GMAIL_ADDRESS, GMAIL_APP_PASS)
            server.sendmail(GMAIL_ADDRESS, [RECIPIENTS[0]], msg.as_string())
        print("  Error notification sent.")
    except Exception as e:
        print(f"  Could not send error email: {e}")

# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    today = date.today()
    yesterday = today - timedelta(days=1)
    date_str = fmt_date_file(today)

    print("=" * 62)
    print(f"  Timiron Cloud Briefing v4 -- {today.strftime('%B %d, %Y')}")
    print(f"  Reporting on: {yesterday}")
    print("=" * 62)

    if not MS_GRAPH_REFRESH_TOKEN:
        print("ERROR: MS_GRAPH_REFRESH_TOKEN not set"); sys.exit(1)
    if not MS_GRAPH_CLIENT_ID:
        print("ERROR: MS_GRAPH_CLIENT_ID not set"); sys.exit(1)
    if not GMAIL_APP_PASS:
        print("ERROR: GMAIL_APP_PASS not set"); sys.exit(1)

    tmpdir = tempfile.mkdtemp(prefix="timiron_")

    try:
        # Step 0: Auth
        print("\n[0] Authenticating with Microsoft Graph...")
        if not get_access_token():
            raise RuntimeError("Could not authenticate with Microsoft Graph. Check MS_GRAPH_REFRESH_TOKEN.")

        # Step 1: Cadiz Ops data
        print("\n[1] Fetching Cadiz Ops data from Outlook...")
        cadiz_data = fetch_cadiz_ops(yesterday)
        print(f"  {len(cadiz_data.get('updates', []))} UPDATE emails found")

        # Step 2: Download load log
        print("\n[2] Downloading Master Load Log...")
        excel_bytes, excel_filename = fetch_load_log_excel(yesterday)
        if not excel_bytes:
            raise RuntimeError("Could not download Master Load Log from email. Check cadiz.ops LOGS emails.")

        # Step 3: Parse
        print("\n[3] Parsing load log...")
        d = parse_load_log(excel_bytes, yesterday)

        # Step 4a: Fetch crew hours from QuickBooks Time
        print("\n[4a] Fetching crew hours from QuickBooks Time...")
        crew_hours, crew_week_label = fetch_qbt_crew_hours()
        if crew_hours:
            print(f"  {len(crew_hours)} employees, WTD: {crew_week_label}")
        else:
            print("  Skipped (no QBT_TOKEN or fetch failed)")

        # Step 4b: Dashboard
        print("\n[4b] Updating Operations Dashboard...")
        dash_tpl = find_template("Operations_Dashboard_MASTER")
        dash_out = os.path.join(tmpdir, f"Timiron_Operations_Dashboard_MASTER_{date_str}.xlsx")
        update_dashboard(dash_tpl, d, dash_out, crew_hours=crew_hours)

        # Step 5: External Report
        print("\n[5] Updating External Report...")
        ext_tpl = find_template("External_Report")
        ext_out = os.path.join(tmpdir, f"Timiron_External_Report_{date_str}.xlsx")
        update_external_report(ext_tpl, d, ext_out)

        # Step 6: Build email
        print("\n[6] Building email...")
        cadiz_section = build_cadiz_section(cadiz_data, d.get('carrier_actuals', {}))
        crew_section = build_crew_hours_html(crew_hours, crew_week_label) if crew_hours else ""
        dash_name = os.path.basename(dash_out)
        ext_name = os.path.basename(ext_out)

        # ATLAS signals (revamp 2026-04-13)
        print("  [6a] Fetching ATLAS pulse...")
        atlas_pulse = fetch_atlas_pulse()
        print(f"  Pulse: integrity={atlas_pulse.get('integrity_passed')}/{atlas_pulse.get('integrity_total')} nyx={atlas_pulse.get('nyx_bot')} bridge={atlas_pulse.get('nyx_bridge')} cadiz={atlas_pulse.get('cadiz_pc')}")

        print("  [6b] Generating ATLAS narrative (Tier 1)...")
        pu_for_narr = d['pump_ute']
        yday_bbls_narr = sum(v['bbls'] for v in pu_for_narr.values())
        yday_trucks_narr = sum(v['loads'] for v in pu_for_narr.values())
        yday_rt_narr = sum(v['runtime'] for v in pu_for_narr.values())
        combined_ute_narr = yday_rt_narr / (PUMP_AVAIL_HRS * 3) * 100 if yday_rt_narr > 0 else 0
        atlas_narrative = generate_atlas_narrative(d, pu_for_narr, yday_bbls_narr, yday_trucks_narr, combined_ute_narr)
        if atlas_narrative:
            print(f"  Narrative ({len(atlas_narrative)} chars): {atlas_narrative[:120]}...")
        else:
            print("  Narrative: empty (skipped)")

        # --- atlas dashboard push ---
        print("  [6c] Pushing ATLAS data to loader dashboard (VERA)...")
        _hero = {
            "yday_bbls": yday_bbls_narr,
            "yday_trucks": yday_trucks_narr,
            "combined_ute_pct": round(combined_ute_narr, 1),
            "mtd_bbls": d["total_bbls"],
            "avg_bbls": d["avg_bbls"],
            "proj_bbls": d["proj_bbls"],
            "proj_trucks": d["proj_trucks"],
            "rail_cap_pct": round(d["rail_cap"] * 100, 1),
        }
        _mkt = _fetch_market_context() if "_fetch_market_context" in globals() else ""
        _push_atlas_to_dashboard(atlas_narrative, atlas_pulse, _hero, _mkt)

        subject = f"\U0001f4ca Timiron Daily Briefing | {today.strftime('%A, %B %d, %Y')}"
        html_body = build_email_html(d, dash_name, ext_name, cadiz_section, crew_section, atlas_pulse, atlas_narrative)
        # NYXP_STATUS_INJECT (C238 follow-up, 2026-04-14 overnight) — append NYXp daily health
        try:
            import sys as _nyxp_sys
            if "/opt/nyx" not in _nyxp_sys.path:
                _nyxp_sys.path.insert(0, "/opt/nyx")
            from nyxp_status import render as _nyxp_render
            _nyxp_text = _nyxp_render()
            _nyxp_section = (
                '<div style="margin:24px 0;padding:14px;background:#0d1117;border:1px solid #30363d;border-radius:8px;font-family:Consolas,monospace;color:#c9d1d9;">' +
                '<div style="color:#58a6ff;font-weight:bold;margin-bottom:8px;">NYXp — Tier Routing &amp; Spend (last 24h)</div>' +
                '<pre style="margin:0;white-space:pre-wrap;font-size:12px;line-height:1.5;">' +
                _nyxp_text.replace('<', '&lt;').replace('>', '&gt;') +
                '</pre></div>'
            )
            if "</body>" in html_body:
                html_body = html_body.replace("</body>", _nyxp_section + "</body>", 1)
            else:
                html_body = html_body + _nyxp_section
        except Exception as _nyxp_e:
            print(f"  NYXp section inject failed: {_nyxp_e}")
        # HERA_COMMERCIAL_INJECT (C246, 2026-04-15) — append HERA commercial intel
        try:
            _hera_section = build_hera_section()
            if _hera_section:
                if "</body>" in html_body:
                    html_body = html_body.replace("</body>", _hera_section + "</body>", 1)
                else:
                    html_body = html_body + _hera_section
                print("  HERA section injected.")
        except Exception as _hera_e:
            print(f"  HERA section inject failed: {_hera_e}")
        # RAIL_PIPELINE_INJECT — append rail pipeline status
        try:
            _rail_section = build_rail_section()
            if _rail_section:
                if "</body>" in html_body:
                    html_body = html_body.replace("</body>", _rail_section + "</body>", 1)
                else:
                    html_body = html_body + _rail_section
                print("  Rail Pipeline section injected.")
        except Exception as _rail_e:
            print(f"  Rail Pipeline section inject failed: {_rail_e}")
        # INVOICE_SUMMARY_INJECT — append financial intelligence
        try:
            _invoice_section = build_invoice_section()
            if _invoice_section:
                if "</body>" in html_body:
                    html_body = html_body.replace("</body>", _invoice_section + "</body>", 1)
                else:
                    html_body = html_body + _invoice_section
                print("  Invoice section injected.")
        except Exception as _inv_e:
            print(f"  Invoice section inject failed: {_inv_e}")
        print(f"  HTML: {len(html_body):,} bytes")

        # Step 7: Send
        print("\n[7] Sending via Gmail...")
        send_via_gmail(subject, html_body, [dash_out, ext_out])

        print("\n" + "=" * 62)
        print(f"  Done. {dash_name} + {ext_name}")
        print("=" * 62)

    except Exception as e:
        error_msg = f"{e}\n\n{traceback.format_exc()}"
        print(f"\nFATAL ERROR: {e}")
        print(traceback.format_exc())
        if GMAIL_APP_PASS:
            send_error_email(error_msg)
        sys.exit(1)
    finally:
        try:
            shutil.rmtree(tmpdir)
        except:
            pass

if __name__ == "__main__":
    main()
